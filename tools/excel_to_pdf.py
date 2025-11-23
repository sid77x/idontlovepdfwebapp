"""Convert Excel to PDF."""
import os
import sys
import streamlit as st
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.file_ops import save_uploaded_file, get_unique_filename, get_file_size_mb
from utils.preview import show_pdf_preview


def excel_to_pdf(input_path: str, output_path: str, orientation: str = "portrait") -> bool:
    """
    Convert Excel to PDF.
    
    Args:
        input_path: Input Excel path
        output_path: Output PDF path
        orientation: "portrait" or "landscape"
    
    Returns:
        bool: Success status
    """
    try:
        # Load Excel workbook
        wb = load_workbook(input_path, data_only=True)
        
        # Set page size and orientation
        if orientation == "landscape":
            pagesize = landscape(letter)
        else:
            pagesize = letter
        
        # Create PDF
        pdf_doc = SimpleDocTemplate(
            output_path,
            pagesize=pagesize,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Process each sheet
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            
            # Add sheet title
            if sheet_idx > 0:
                elements.append(PageBreak())
            
            title = Paragraph(f"<b>{sheet_name}</b>", styles['Heading1'])
            elements.append(title)
            elements.append(Spacer(1, 0.2 * inch))
            
            # Extract data from sheet
            data = []
            for row in ws.iter_rows(values_only=True):
                # Convert None to empty string and handle all types
                row_data = []
                for cell in row:
                    if cell is None:
                        row_data.append('')
                    else:
                        row_data.append(str(cell))
                
                # Skip completely empty rows
                if any(cell for cell in row_data):
                    data.append(row_data)
            
            if not data:
                elements.append(Paragraph("(Empty sheet)", styles['Normal']))
                continue
            
            # Calculate column widths (simple approach)
            if data:
                num_cols = len(data[0]) if data else 0
                available_width = pagesize[0] - 60  # Total width minus margins
                col_width = available_width / num_cols if num_cols > 0 else 100
                col_widths = [col_width] * num_cols
            else:
                col_widths = None
            
            # Create table
            table = Table(data, colWidths=col_widths, repeatRows=1)
            
            # Style the table
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ])
            table.setStyle(table_style)
            
            elements.append(table)
            elements.append(Spacer(1, 0.3 * inch))
        
        # Build PDF
        pdf_doc.build(elements)
        wb.close()
        
        return True
        
    except Exception as e:
        st.error(f"Error converting Excel to PDF: {str(e)}")
        return False


def render_excel_to_pdf_tool():
    """Render the Excel to PDF tool interface."""
    st.title("📊 Excel to PDF")
    st.write("Convert your Excel spreadsheets to PDF format")
    
    # File upload
    uploaded_file = st.file_uploader(
        "Choose an Excel file",
        type=['xlsx', 'xls'],
        help="Select the Excel file you want to convert to PDF"
    )
    
    if uploaded_file:
        # Display file info
        file_size = get_file_size_mb(uploaded_file)
        st.info(f"📊 **{uploaded_file.name}** ({file_size:.2f} MB)")
        
        # Save uploaded file
        input_path = save_uploaded_file(uploaded_file, "excel_to_pdf")
        
        if input_path:
            try:
                # Load workbook to get info
                wb = load_workbook(input_path, read_only=True, data_only=True)
                num_sheets = len(wb.sheetnames)
                sheet_names = wb.sheetnames
                wb.close()
                
                st.success(f"✅ Excel file loaded: {num_sheets} sheet(s)")
                
                with st.expander("📋 Sheet names"):
                    for idx, name in enumerate(sheet_names, 1):
                        st.write(f"{idx}. {name}")
                
                # Conversion settings
                st.write("### ⚙️ Conversion Settings")
                
                orientation = st.radio(
                    "Page Orientation",
                    ["portrait", "landscape"],
                    horizontal=True,
                    help="Landscape may be better for wide spreadsheets"
                )
                
                st.info("💡 Each sheet will be converted to a separate page in the PDF.")
                
                # Convert button
                if st.button("🔄 Convert to PDF", type="primary", use_container_width=True):
                    with st.spinner("Converting Excel to PDF..."):
                        # Create output path
                        output_dir = os.path.join("output", "excel_to_pdf")
                        os.makedirs(output_dir, exist_ok=True)
                        output_path = os.path.join(output_dir, get_unique_filename("converted.pdf"))
                        
                        # Convert Excel to PDF
                        success = excel_to_pdf(input_path, output_path, orientation)
                        
                        if success:
                            st.success(f"✅ Successfully converted Excel to PDF!")
                            
                            # Show preview
                            st.write("### 🔍 Preview")
                            try:
                                with open(output_path, "rb") as f:
                                    preview_data = f.read()
                                
                                show_pdf_preview(preview_data, height=1000)
                            except Exception as e:
                                st.warning(f"Preview unavailable: {str(e)}")
                            
                            # Show file info
                            output_size = os.path.getsize(output_path) / (1024 * 1024)
                            st.info(f"📄 Output PDF size: {output_size:.2f} MB")
                            
                            # Download button
                            with open(output_path, "rb") as f:
                                st.download_button(
                                    label="⬇️ Download PDF",
                                    data=f,
                                    file_name="converted.pdf",
                                    mime="application/pdf",
                                    use_container_width=True
                                )
                
            except Exception as e:
                st.error(f"❌ Error processing Excel file: {str(e)}")
    
    else:
        # Help section
        st.write("### 📖 How to use")
        st.markdown("""
        1. **Upload** your Excel file (.xlsx or .xls)
        2. **Choose** page orientation
        3. **Click** Convert to PDF
        4. **Download** your PDF file
        
        **Features:**
        - ✅ All sheets converted
        - ✅ Tables with styling
        - ✅ Grid lines preserved
        - ✅ Automatic column sizing
        
        **Best for:**
        - Spreadsheets with data tables
        - Reports and summaries
        - Sharing data in read-only format
        
        **Tips:**
        - Use landscape orientation for wide sheets
        - Each sheet becomes a new page
        - Formulas are converted to values
        - Charts and images may not be included
        """)


if __name__ == "__main__":
    render_excel_to_pdf_tool()
