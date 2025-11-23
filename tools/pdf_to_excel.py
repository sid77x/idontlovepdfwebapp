"""Convert PDF tables to Excel."""
import os
import sys
import streamlit as st
import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.file_ops import save_uploaded_file, get_unique_filename, get_file_size_mb


def pdf_to_excel(input_path: str, output_path: str) -> bool:
    """
    Convert PDF tables to Excel.
    
    Args:
        input_path: Input PDF path
        output_path: Output Excel path
    
    Returns:
        bool: Success status
    """
    try:
        # Open PDF
        pdf_document = fitz.open(input_path)
        
        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Process each page
        for page_num in range(len(pdf_document)):
            page = pdf_document[page_num]
            
            # Create a sheet for this page
            ws = wb.create_sheet(title=f"Page {page_num + 1}")
            
            # Extract text
            text = page.get_text()
            
            # Split into lines
            lines = text.split('\n')
            
            # Add text to Excel (simple line-by-line)
            for row_idx, line in enumerate(lines, start=1):
                if line.strip():
                    # Try to split by common delimiters
                    if '\t' in line:
                        cells = line.split('\t')
                    elif '  ' in line:  # Multiple spaces
                        cells = [cell.strip() for cell in line.split('  ') if cell.strip()]
                    else:
                        cells = [line]
                    
                    # Add cells to row
                    for col_idx, cell in enumerate(cells, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=cell)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Style header row (first row)
            if ws.max_row > 0:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
        
        # Save workbook
        wb.save(output_path)
        pdf_document.close()
        
        return True
        
    except Exception as e:
        st.error(f"Error converting PDF to Excel: {str(e)}")
        return False


def render_pdf_to_excel_tool():
    """Render the PDF to Excel tool interface."""
    st.title("📊 PDF to Excel")
    st.write("Extract tables and text from PDF to Excel format")
    
    # File upload
    uploaded_file = st.file_uploader(
        "Choose a PDF file",
        type=['pdf'],
        help="Select the PDF file you want to convert to Excel"
    )
    
    if uploaded_file:
        # Display file info
        file_size = get_file_size_mb(uploaded_file)
        st.info(f"📄 **{uploaded_file.name}** ({file_size:.2f} MB)")
        
        # Save uploaded file
        input_path = save_uploaded_file(uploaded_file, "pdf_to_excel")
        
        if input_path:
            try:
                # Get page count
                pdf_doc = fitz.open(input_path)
                total_pages = len(pdf_doc)
                pdf_doc.close()
                
                st.success(f"✅ PDF loaded: {total_pages} pages")
                
                st.warning("⚠️ **Important:** This tool provides basic table extraction. For complex tables, consider specialized OCR tools.")
                st.info("💡 Each PDF page will be converted to a separate Excel sheet.")
                
                # Convert button
                if st.button("🔄 Convert to Excel", type="primary", use_container_width=True):
                    with st.spinner("Converting PDF to Excel..."):
                        # Create output path
                        output_dir = os.path.join("output", "pdf_to_excel")
                        os.makedirs(output_dir, exist_ok=True)
                        output_path = os.path.join(output_dir, get_unique_filename("converted.xlsx"))
                        
                        # Convert PDF to Excel
                        success = pdf_to_excel(input_path, output_path)
                        
                        if success:
                            st.success(f"✅ Successfully converted PDF to Excel!")
                            
                            # Show file info
                            output_size = os.path.getsize(output_path) / (1024 * 1024)
                            st.info(f"📊 Output Excel size: {output_size:.2f} MB • {total_pages} sheets")
                            
                            # Download button
                            with open(output_path, "rb") as f:
                                st.download_button(
                                    label="⬇️ Download Excel",
                                    data=f,
                                    file_name="converted.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True
                                )
                            
                            st.success("💡 **Tip:** Open in Excel or Google Sheets to review and edit the data.")
                
            except Exception as e:
                st.error(f"❌ Error processing PDF: {str(e)}")
    
    else:
        # Help section
        st.write("### 📖 How to use")
        st.markdown("""
        1. **Upload** your PDF file
        2. **Click** Convert to Excel
        3. **Download** your Excel file
        
        **What to expect:**
        - Each page becomes a separate sheet
        - Text is extracted line by line
        - Automatic column width adjustment
        - Header row styling
        
        **Best for:**
        - PDFs with simple tables
        - Data extraction
        - Reports with tabular data
        
        **Limitations:**
        - ⚠️ Complex table structures may not be preserved
        - ⚠️ Merged cells are not detected
        - ⚠️ Images are not included
        - ⚠️ Best results with text-based PDFs
        
        **Tip:** For scanned documents, consider using OCR tools first.
        """)


if __name__ == "__main__":
    render_pdf_to_excel_tool()
